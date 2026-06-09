import json
import sys
import threading
import time
import urllib.request
from io import BytesIO
from pathlib import Path
from uuid import uuid4

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


def make_workspace_tmp():
    path = Path("annotations_test_tmp") / uuid4().hex
    path.mkdir(parents=True, exist_ok=False)
    return path


def write_jsonl(path, rows):
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def make_test_image(path, size):
    from PIL import Image

    path.parent.mkdir(parents=True, exist_ok=True)
    Image.new("RGB", size, color=(20, 120, 220)).save(path, format="JPEG")


def test_normalize_annotation_tags_stores_known_dimensions_as_single_values():
    from web.annotations.app import normalize_annotation_tags
    from web.annotations.label_options import LABEL_OPTION_GROUPS

    group = LABEL_OPTION_GROUPS[0]
    dimension = group["dimensions"][1]
    group_name = group["name"]
    dimension_name = dimension["name"]
    first_option, second_option = dimension["options"][:2]

    normalized = normalize_annotation_tags({
        group_name: {
            dimension_name: [first_option, second_option],
            "custom": ["keep", "as-list"],
        },
        "other": {"value": ["also", "kept"]},
    })

    assert normalized[group_name][dimension_name] == first_option
    assert normalized[group_name]["custom"] == ["keep", "as-list"]
    assert normalized["other"]["value"] == ["also", "kept"]


def test_create_task_splits_jsonl_into_ordered_subtasks():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    root = tmp_path / "images"
    root.mkdir()
    annotation_dir = tmp_path / "labels"
    jsonl_path = tmp_path / "data.jsonl"
    rows = [
        {"src_image": f"src/{idx}.jpg", "dst_image": f"dst/{idx}.jpg", "labels": {"idx": "ignored"}}
        for idx in range(5)
    ]
    write_jsonl(jsonl_path, rows)
    write_json(annotation_dir / "src" / "3.json", {"labels": {"菜品种类": "中餐"}})
    write_json(annotation_dir / "dst" / "3.json", {"美学评分": 4})

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(root), str(jsonl_path), chunk_size=2, annotation_dir=str(annotation_dir))

    assert task["name"] == "food"
    assert task["annotation_dir"] == str(annotation_dir)
    assert [subtask["item_indexes"] for subtask in task["subtasks"]] == [[0, 1], [2, 3], [4]]
    assert task["items"][3]["labels"] == {"输入图": {"菜品种类": "中餐"}, "输出图": {"美学评分": 4}}
    assert task["items"][0]["labels"] == {}
    assert "tags" not in task["items"][3]

    state_data = json.loads((tmp_path / "state.json").read_text(encoding="utf-8"))
    assert "items" not in state_data["tasks"][0]
    assert "annotations" not in state_data["tasks"][0]
    assert Path(state_data["tasks"][0]["data_dir"]).exists()
    data_dir = Path(state_data["tasks"][0]["data_dir"])
    assert state_data["tasks"][0]["items_storage"] == "chunks"
    assert state_data["tasks"][0]["annotations_storage"] == "items"
    assert len(list((data_dir / "items").glob("*.json"))) == 3
    assert not (data_dir / "items.json").exists()
    assert not (data_dir / "annotations.json").exists()


def test_create_task_can_shuffle_items_before_splitting(monkeypatch):
    from web.annotations import app as annotations_module
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    rows = [{"src_image": f"src/{idx}.jpg", "dst_image": f"dst/{idx}.jpg"} for idx in range(5)]
    write_jsonl(jsonl_path, rows)
    monkeypatch.setattr(annotations_module.random, "shuffle", lambda items: items.reverse())

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=2, shuffle_items=True)

    assert task["shuffle_items"] is True
    assert [item["src_image"] for item in task["items"]] == [
        "src/4.jpg",
        "src/3.jpg",
        "src/2.jpg",
        "src/1.jpg",
        "src/0.jpg",
    ]
    assert [subtask["item_indexes"] for subtask in task["subtasks"]] == [[0, 1], [2, 3], [4]]
    assert [item["src_image"] for item in store._read_subtask_items(task, task["subtasks"][0])] == [
        "src/4.jpg",
        "src/3.jpg",
    ]


def test_create_task_job_reports_progress_and_summary():
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [{"src_image": f"a{idx}.jpg", "dst_image": f"b{idx}.jpg"} for idx in range(3)],
    )

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        client = annotations_app.app.test_client()
        response = client.post(
            "/api/tasks/jobs",
            json={
                "name": "food",
                "root_dir": str(tmp_path),
                "jsonl_path": str(jsonl_path),
                "chunk_size": 2,
            },
        )
        assert response.status_code == 202
        job_id = response.get_json()["job"]["id"]

        for _ in range(50):
            job_response = client.get(f"/api/tasks/jobs/{job_id}")
            job = job_response.get_json()["job"]
            if job["status"] == "completed":
                break
            time.sleep(0.02)
        else:
            raise AssertionError("create task job did not complete")

        assert job["progress"] == 100
        assert job["task"]["name"] == "food"
        assert job["task"]["item_count"] == 3
    finally:
        annotations_app.store = old_store


def test_annotation_messages_and_json_responses_keep_utf8_chinese_readable():
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore, load_jsonl

    source = (PROJECT_ROOT / "web" / "annotations" / "app.py").read_text(encoding="utf-8")
    for garbled_marker in ["绗", "琛", "鏍", "姝", "瀛", "浠"]:
        assert garbled_marker not in source

    tmp_path = make_workspace_tmp()
    bad_jsonl_path = tmp_path / "坏数据.jsonl"
    write_jsonl(bad_jsonl_path, [{"src_image": "原图/a.jpg"}])
    try:
        load_jsonl(str(bad_jsonl_path))
    except ValueError as exc:
        assert str(exc) == "第 1 行缺少 src_image 或 dst_image"
    else:
        raise AssertionError("load_jsonl should reject rows missing dst_image")

    good_jsonl_path = tmp_path / "好数据.jsonl"
    write_jsonl(good_jsonl_path, [{"src_image": "原图/a.jpg", "dst_image": "目标图/a.jpg"}])
    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("中文任务", str(tmp_path), str(good_jsonl_path), chunk_size=1)
        client = annotations_app.app.test_client()
        response = client.get(f"/api/tasks/{task['id']}/download?format=xml")

        assert response.status_code == 400
        assert response.content_type == "application/json"
        assert response.data.decode("utf-8") == '{"error":"不支持的下载格式"}\n'
    finally:
        annotations_app.store = old_store


def test_store_migrates_legacy_inline_task_data_to_split_files():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    task_id = "legacy-task"
    legacy_state = {
        "tasks": [
            {
                "id": task_id,
                "name": "legacy",
                "root_dir": str(tmp_path),
                "jsonl_path": str(tmp_path / "data.jsonl"),
                "annotation_dir": "",
                "chunk_size": 1,
                "created_at": 1,
                "items": [{"src_image": "a.jpg", "dst_image": "b.jpg", "labels": {}}],
                "subtasks": [
                    {
                        "id": "subtask",
                        "index": 1,
                        "item_indexes": [0],
                        "assigned_to": None,
                        "assigned_at": None,
                        "completed_at": None,
                        "completed_count": 0,
                    }
                ],
                "annotations": {"0": {"mos": 4, "tags": {}}},
            }
        ]
    }
    write_json(tmp_path / "state.json", legacy_state)

    store = AnnotationStore(tmp_path / "state.json")
    state_data = json.loads((tmp_path / "state.json").read_text(encoding="utf-8"))
    migrated = state_data["tasks"][0]

    assert "items" not in migrated
    assert "annotations" not in migrated
    assert migrated["item_count"] == 1
    assert migrated["annotation_count"] == 1
    assert migrated["items_storage"] == "chunks"
    assert migrated["annotations_storage"] == "items"
    assert store.get_task(task_id)["items"][0]["src_image"] == "a.jpg"
    assert (Path(migrated["data_dir"]) / "items" / "subtask.json").exists()
    assert (Path(migrated["data_dir"]) / "annotations" / "0.json").exists()


def test_subtask_payload_uses_image_label_json_and_hides_unseen_label_dimensions():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    annotation_dir = tmp_path / "标注结果"
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [
            {
                "src_image": "原始图片/a.jpg",
                "dst_image": "生成图片/b.jpg",
                "labels": {"输入图": {"光线": "自然光"}},
                "tags": {"输入图": {"光线": "自然光"}},
            },
            {
                "src_image": "原始图片/c.jpg",
                "dst_image": "生成图片/d.jpg",
            },
        ],
    )
    write_json(annotation_dir / "原始图片" / "a.json", {"labels": {"菜品种类": "中餐"}})
    write_json(annotation_dir / "生成图片" / "b.json", {"labels": {"输出图": {"美学评分": 4}}})
    write_json(annotation_dir / "原始图片" / "c.json", {"菜品种类": "西餐"})

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=2, annotation_dir=str(annotation_dir))
    subtask = store.assign_subtask(task["id"], "alice")

    assert subtask["items"][0]["labels"] == {"输入图": {"菜品种类": "中餐"}, "输出图": {"美学评分": 4}}
    assert subtask["items"][0]["tags"] == {"输入图": {"菜品种类": "中餐"}, "输出图": {"美学评分": 4}}

    visible = {
        (group["name"], dimension["name"])
        for group in subtask["label_options"]
        for dimension in group["dimensions"]
    }
    assert ("输入图", "菜品种类") in visible
    assert ("输出图", "美学评分") in visible
    assert ("输入图", "光线") not in visible


def test_read_image_description_prompts_from_ai_label_json():
    from web.annotations.app import read_image_description_prompts

    tmp_path = make_workspace_tmp()
    root = tmp_path / "images"
    annotation_dir = tmp_path / "labels"
    content = {
        "adjustment_aigc": "AIGC text",
        "adjustment_user": "User prompt",
        "thinking": "Reasoning text",
    }
    write_json(
        annotation_dir / "dst" / "sample.json",
        {
            "description": {
                "conversation_history": [
                    {},
                    {},
                    {},
                    {},
                    {},
                    {"content": json.dumps(content, ensure_ascii=False)},
                ]
            }
        },
    )

    prompts = read_image_description_prompts(root, annotation_dir, "dst/sample.jpg")

    assert prompts == content


def test_read_image_description_prompts_accepts_fenced_history_answer():
    from web.annotations.app import read_image_description_prompts

    tmp_path = make_workspace_tmp()
    root = tmp_path / "images"
    annotation_dir = tmp_path / "labels"
    content = {
        "adjustment_aigc": "AIGC text from fenced answer",
        "adjustment_user": "User prompt from fenced answer",
        "thinking": "Reasoning text from fenced answer",
    }
    write_json(
        annotation_dir / "dst" / "sample.json",
        {
            "description": {
                "conversation_history": [
                    {"role": "system", "content": "schema"},
                    {"role": "user", "content": [{"type": "text", "text": "question"}]},
                    {
                        "role": "assistant",
                        "content": "```json\n" + json.dumps(content, ensure_ascii=False) + "\n```",
                    },
                ]
            }
        },
    )

    prompts = read_image_description_prompts(root, annotation_dir, "dst/sample.jpg")

    assert prompts == content


def test_visualization_results_api_includes_description_prompts():
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    root = tmp_path / "images"
    annotation_dir = tmp_path / "labels"
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [
            {"src_image": "src/a.jpg", "dst_image": "dst/b.jpg"},
            {"src_image": "src/c.jpg", "dst_image": "dst/d.jpg"},
        ],
    )
    write_json(
        annotation_dir / "dst" / "b.json",
        {
            "description": {
                "conversation_history": [
                    {},
                    {},
                    {},
                    {},
                    {},
                    {
                        "content": json.dumps(
                            {
                                "adjustment_aigc": "AIGC caption",
                                "adjustment_user": "User instruction",
                                "thinking": "Model thoughts",
                            },
                            ensure_ascii=False,
                        )
                    },
                ]
            },
            "labels": {"output": {"quality": "good"}},
        },
    )
    write_json(
        annotation_dir / "dst" / "d.json",
        {
            "description": {
                "conversation_history": [
                    {},
                    {},
                    {},
                    {},
                    {},
                    {
                        "content": json.dumps(
                            {
                                "adjustment_aigc": "Unscored AIGC caption",
                                "adjustment_user": "Unscored user instruction",
                                "thinking": "Unscored model thoughts",
                            },
                            ensure_ascii=False,
                        )
                    },
                ]
            },
            "labels": {"output": {"quality": "unscored-label"}},
        },
    )

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("food", str(root), str(jsonl_path), chunk_size=2, annotation_dir=str(annotation_dir))
        subtask = annotations_app.store.assign_subtask(task["id"], "alice")
        annotations_app.store.save_annotation(task["id"], subtask["id"], 0, "alice", 5, {"score": "ok"})
        client = annotations_app.app.test_client()

        response = client.get(f"/api/tasks/{task['id']}/visualization-results")

        assert response.status_code == 200
        payload = response.get_json()
        assert payload["total"] == 2
        results = payload["results"]
        assert [item["item_index"] for item in results] == [0, 1]
        item = results[0]
        assert item["item_index"] == 0
        assert item["mos"] == 5
        assert item["tags"] == {"score": "ok"}
        assert item["description_prompts"] == {
            "adjustment_aigc": "AIGC caption",
            "adjustment_user": "User instruction",
            "thinking": "Model thoughts",
        }
        unscored = results[1]
        assert unscored["mos"] is None
        assert unscored["username"] is None
        assert unscored["tags"] == {"输出图": {"output": {"quality": "unscored-label"}}}
        assert unscored["description_prompts"] == {
            "adjustment_aigc": "Unscored AIGC caption",
            "adjustment_user": "Unscored user instruction",
            "thinking": "Unscored model thoughts",
        }

        paged_response = client.get(f"/api/tasks/{task['id']}/visualization-results?page=1&limit=1")
        assert paged_response.status_code == 200
        paged_payload = paged_response.get_json()
        assert paged_payload["total"] == 2
        assert [item["item_index"] for item in paged_payload["results"]] == [1]
    finally:
        annotations_app.store = old_store


def test_assign_subtask_is_exclusive_and_reuses_users_active_subtask():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [{"src_image": f"a{idx}.jpg", "dst_image": f"b{idx}.jpg"} for idx in range(4)],
    )

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=2)

    alice_first = store.assign_subtask(task["id"], "alice")
    alice_second = store.assign_subtask(task["id"], "alice")
    bob = store.assign_subtask(task["id"], "bob")

    assert alice_first["id"] == alice_second["id"]
    assert bob["id"] != alice_first["id"]
    assert store.assign_subtask(task["id"], "carol") is None


def test_abandon_subtask_deletes_its_annotations_and_releases_assignment():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [{"src_image": f"src/{idx}.jpg", "dst_image": f"dst/{idx}.jpg"} for idx in range(4)],
    )

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=2)
    alice_subtask = store.assign_subtask(task["id"], "alice")
    bob_subtask = store.assign_subtask(task["id"], "bob")
    alice_indexes = alice_subtask["item_indexes"]
    bob_index = bob_subtask["item_indexes"][0]
    store.save_annotation(task["id"], alice_subtask["id"], alice_indexes[0], "alice", 2, {"bad": True})
    store.save_annotation(task["id"], alice_subtask["id"], alice_indexes[1], "alice", 5, {"good": True})
    store.save_annotation(task["id"], bob_subtask["id"], bob_index, "bob", 4, {"keep": True})

    result = store.abandon_subtask(task["id"], alice_subtask["id"], "alice")

    assert result["deleted_count"] == 2
    assert result["subtask"]["assigned_to"] is None
    assert result["subtask"]["assigned_at"] is None
    assert result["subtask"]["completed_at"] is None
    assert result["subtask"]["completed_count"] == 0
    refreshed = store.get_task(task["id"])
    assert refreshed["annotation_count"] == 1
    assert store.get_results(task["id"], threshold=1)[0]["item_index"] == bob_index
    data_dir = Path(refreshed["data_dir"])
    assert not (data_dir / "annotations" / f"{alice_indexes[0]}.json").exists()
    assert not (data_dir / "annotations" / f"{alice_indexes[1]}.json").exists()
    assert (data_dir / "annotations" / f"{bob_index}.json").exists()
    reassigned = store.assign_subtask(task["id"], "carol")
    assert reassigned["id"] == alice_subtask["id"]


def test_abandon_subtask_api_requires_owner_and_returns_summary():
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [{"src_image": f"src/{idx}.jpg", "dst_image": f"dst/{idx}.jpg"} for idx in range(2)],
    )

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=2)
        subtask = annotations_app.store.assign_subtask(task["id"], "alice")
        annotations_app.store.save_annotation(task["id"], subtask["id"], 0, "alice", 2, {"bad": True})
        client = annotations_app.app.test_client()

        forbidden = client.delete(
            f"/api/tasks/{task['id']}/subtasks/{subtask['id']}",
            json={"username": "bob"},
        )
        response = client.delete(
            f"/api/tasks/{task['id']}/subtasks/{subtask['id']}",
            json={"username": "alice"},
        )

        assert forbidden.status_code == 403
        assert response.status_code == 200
        payload = response.get_json()
        assert payload["deleted_count"] == 1
        assert payload["subtask"]["assigned_to"] is None
    finally:
        annotations_app.store = old_store


def test_save_annotation_marks_progress_and_results_filter_by_threshold():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    rows = [
        {
            "src_image": f"src/{idx}.jpg",
            "dst_image": f"dst/{idx}.jpg",
            "tags": {"输出图": {"美学评分": idx}},
        }
        for idx in range(3)
    ]
    write_jsonl(jsonl_path, rows)

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=3)
    subtask = store.assign_subtask(task["id"], "alice")

    store.save_annotation(
        task["id"],
        subtask["id"],
        item_index=0,
        username="alice",
        mos=3,
        tags={"bad": True},
    )
    store.save_annotation(
        task["id"],
        subtask["id"],
        item_index=1,
        username="alice",
        mos=4,
        tags={"good": True},
    )

    refreshed = store.get_task(task["id"])
    assert refreshed["subtasks"][0]["completed_count"] == 2
    data_dir = Path(refreshed["data_dir"])
    assert (data_dir / "annotations" / "0.json").exists()
    assert (data_dir / "annotations" / "1.json").exists()
    assert not (data_dir / "annotations.json").exists()

    results = store.get_results(task["id"], threshold=4)
    assert [result["item_index"] for result in results] == [1]
    assert results[0]["username"] == "alice"
    assert results[0]["mos"] == 4
    assert results[0]["tags"] == {"good": True}


def test_task_summary_recomputes_annotation_count_from_annotation_files():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [{"src_image": f"src/{idx}.jpg", "dst_image": f"dst/{idx}.jpg"} for idx in range(3)],
    )

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=3)
    subtask = store.assign_subtask(task["id"], "alice")
    store.save_annotation(task["id"], subtask["id"], item_index=0, username="alice", mos=3, tags={})
    store.save_annotation(task["id"], subtask["id"], item_index=1, username="alice", mos=4, tags={})

    state_data = json.loads((tmp_path / "state.json").read_text(encoding="utf-8"))
    state_data["tasks"][0]["annotation_count"] = 99
    write_json(tmp_path / "state.json", state_data)

    summary = store.list_tasks()[0]

    assert summary["annotation_count"] == len(store.get_results(task["id"], threshold=1)) == 2


def test_backup_annotations_writes_completed_results_under_annotation_dir_backup():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    annotation_dir = tmp_path / "labels"
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [{"src_image": f"src/{idx}.jpg", "dst_image": f"dst/{idx}.jpg"} for idx in range(2)],
    )

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=2, annotation_dir=str(annotation_dir))
    subtask = store.assign_subtask(task["id"], "alice")
    store.save_annotation(task["id"], subtask["id"], item_index=0, username="alice", mos=4, tags={"ok": True})

    result = store.backup_annotations()

    backup_root = Path(str(annotation_dir) + "_bak")
    backup_file = backup_root / task["id"] / "annotations" / "0.json"
    manifest_file = backup_root / task["id"] / "manifest.json"
    assert result["task_count"] == 1
    assert result["annotation_count"] == 1
    assert backup_file.exists()
    assert json.loads(backup_file.read_text(encoding="utf-8"))["mos"] == 4
    assert json.loads(manifest_file.read_text(encoding="utf-8"))["source_annotation_dir"] == str(annotation_dir)


def test_results_filter_by_mos_annotator_and_label_dimensions():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [{"src_image": f"src/{idx}.jpg", "dst_image": f"dst/{idx}.jpg"} for idx in range(4)],
    )

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=2)
    alice_subtask = store.assign_subtask(task["id"], "alice")
    bob_subtask = store.assign_subtask(task["id"], "bob")
    subtask_by_item = {
        item_index: (alice_subtask["id"], "alice")
        for item_index in alice_subtask["item_indexes"]
    }
    subtask_by_item.update(
        {
            item_index: (bob_subtask["id"], "bob")
            for item_index in bob_subtask["item_indexes"]
        }
    )
    annotations = [
        (5, {"输出图": {"菜品种类": ["中餐"], "光线": ["自然光"]}}),
        (5, {"输出图": {"菜品种类": ["西餐"], "光线": ["顶灯"]}}),
        (5, {"输出图": {"菜品种类": ["中餐"], "光线": ["自然光"]}}),
        (4, {"输出图": {"菜品种类": ["西餐"], "光线": ["自然光"]}}),
    ]
    for item_index, (mos, tags) in enumerate(annotations):
        subtask_id, username = subtask_by_item[item_index]
        store.save_annotation(task["id"], subtask_id, item_index, username, mos, tags)

    filters = {
        "mos": [5],
        "annotators": [subtask_by_item[0][1]],
        "labels": [
            {"path": ["输出图", "菜品种类"], "values": ["中餐", "西餐"]},
            {"path": ["输出图", "光线"], "values": ["自然光"]},
        ],
    }

    results = store.get_results(task["id"], filters=filters)

    assert [result["item_index"] for result in results] == [0]


def test_results_filter_supports_numeric_label_ranges():
    from web.annotations.app import AnnotationStore
    from web.annotations.label_options import LABEL_OPTION_GROUPS

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [{"src_image": f"src/{idx}.jpg", "dst_image": f"dst/{idx}.jpg"} for idx in range(3)],
    )
    angle_path = [LABEL_OPTION_GROUPS[0]["name"], LABEL_OPTION_GROUPS[0]["dimensions"][0]["name"]]

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=3)
    subtask = store.assign_subtask(task["id"], "alice")
    for item_index, angle in enumerate([0, 45, 75]):
        store.save_annotation(
            task["id"],
            subtask["id"],
            item_index=item_index,
            username="alice",
            mos=5,
            tags={angle_path[0]: {angle_path[1]: angle}},
        )

    results = store.get_results(
        task["id"],
        threshold=1,
        filters={"labels": [{"path": angle_path, "ranges": [{"min": 0, "max": 60}]}]},
    )

    assert [result["item_index"] for result in results] == [0, 1]


def test_pair_label_options_are_global_and_filterable_across_tasks():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    first_jsonl = tmp_path / "first.jsonl"
    second_jsonl = tmp_path / "second.jsonl"
    write_jsonl(first_jsonl, [{"src_image": "src/a.jpg", "dst_image": "dst/a.jpg"}])
    write_jsonl(second_jsonl, [{"src_image": "src/b.jpg", "dst_image": "dst/b.jpg"}])

    store = AnnotationStore(tmp_path / "state.json")
    first_task = store.create_task("first", str(tmp_path), str(first_jsonl), chunk_size=1)
    second_task = store.create_task("second", str(tmp_path), str(second_jsonl), chunk_size=1)

    options = store.add_pair_label_option("问题标签", "主体不清晰")

    assert {"name": "问题标签", "options": ["主体不清晰"]} in options["dimensions"]
    assert store.assign_subtask(first_task["id"], "alice")["pair_label_options"] == options
    second_subtask = store.assign_subtask(second_task["id"], "bob")
    assert second_subtask["pair_label_options"] == options

    store.save_annotation(
        second_task["id"],
        second_subtask["id"],
        item_index=0,
        username="bob",
        mos=5,
        tags={"Pair": {"问题标签": "主体不清晰", "优势标签": "氛围自然"}},
    )

    filter_options = store.get_result_filter_options(second_task["id"])
    pair_group = next(group for group in filter_options["label_options"] if group["name"] == "Pair")
    assert pair_group["dimensions"] == [
        {"name": "问题标签", "options": ["主体不清晰"]},
        {"name": "优势标签", "options": ["氛围自然"]},
    ]

    results = store.get_results(
        second_task["id"],
        threshold=1,
        filters={"labels": [{"path": ["Pair", "问题标签"], "values": ["主体不清晰"]}]},
    )
    assert [result["item_index"] for result in results] == [0]

    statistics = store.get_statistics(
        second_task["id"],
        filters={"labels": [{"path": ["Pair", "优势标签"], "values": ["氛围自然"]}]},
    )
    assert statistics["total"] == 1
    assert {"type": "label", "path": ["Pair", "问题标签"], "label": "Pair / 问题标签"} in statistics["available_dimensions"]
    assert {"type": "label", "path": ["Pair", "优势标签"], "label": "Pair / 优势标签"} in statistics["available_dimensions"]


def test_pair_label_option_api_adds_global_option():
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        client = annotations_app.app.test_client()

        response = client.post(
            "/api/pair-label-options",
            json={"dimension": "优势标签", "label": "构图稳定"},
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload["pair_label_options"] == {
            "name": "Pair",
            "dimensions": [
                {"name": "问题标签", "options": []},
                {"name": "优势标签", "options": ["构图稳定"]},
            ],
        }
    finally:
        annotations_app.store = old_store


def test_quality_check_updates_final_result_and_keeps_history():
    from web.annotations.app import AnnotationStore
    from web.annotations.label_options import LABEL_OPTION_GROUPS

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "a.jpg", "dst_image": "b.jpg"}])
    input_group = LABEL_OPTION_GROUPS[0]["name"]
    angle_name = LABEL_OPTION_GROUPS[0]["dimensions"][0]["name"]

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
    subtask = store.assign_subtask(task["id"], "alice")
    store.save_annotation(task["id"], subtask["id"], 0, "alice", 4, {input_group: {angle_name: 10}})

    annotation = store.save_quality_check(
        task["id"],
        0,
        "bob",
        5,
        {input_group: {angle_name: 45}},
    )

    assert annotation["username"] == "alice"
    assert annotation["mos"] == 5
    assert annotation["tags"] == {input_group: {angle_name: 45}}
    assert annotation["qc_reviewers"] == ["bob"]
    assert annotation["qc_history"][0]["username"] == "bob"
    assert annotation["qc_history"][0]["before"]["mos"] == 4
    assert annotation["qc_history"][0]["after"]["tags"] == {input_group: {angle_name: 45}}

    results = store.get_results(task["id"], threshold=1)
    assert results[0]["mos"] == 5
    assert results[0]["tags"] == {input_group: {angle_name: 45}}
    assert results[0]["qc_reviewers"] == ["bob"]


def test_quality_check_tracks_multiple_reviewers_but_exports_final_result_only():
    from web.annotations.app import AnnotationStore, export_row

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "a.jpg", "dst_image": "b.jpg"}])

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
    subtask = store.assign_subtask(task["id"], "alice")
    store.save_annotation(task["id"], subtask["id"], 0, "alice", 3, {"stage": "annotated"})
    store.save_quality_check(task["id"], 0, "bob", 4, {"stage": "bob"})
    store.save_quality_check(task["id"], 0, "carol", 5, {"stage": "carol"})
    store.save_quality_check(task["id"], 0, "alice", 4, {"stage": "self-check"})

    result = store.get_results(task["id"], threshold=1)[0]

    assert result["mos"] == 4
    assert result["tags"] == {"stage": "self-check"}
    assert result["qc_reviewers"] == ["bob", "carol"]
    assert [record["username"] for record in result["qc_history"]] == ["bob", "carol", "alice"]

    exported = export_row(result)
    assert exported["mos"] == 4
    assert exported["tags"] == {"stage": "self-check"}
    assert "qc_history" not in exported


def test_quality_check_undo_only_removes_current_users_latest_edit_and_replays_later_records():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "a.jpg", "dst_image": "b.jpg"}])

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
    subtask = store.assign_subtask(task["id"], "alice")
    store.save_annotation(task["id"], subtask["id"], 0, "alice", 3, {"stage": "annotated"})
    store.save_quality_check(task["id"], 0, "bob", 4, {"stage": "bob"})
    store.save_quality_check(task["id"], 0, "carol", 5, {"stage": "carol"})

    annotation = store.undo_quality_check(task["id"], 0, "bob")

    assert annotation["mos"] == 5
    assert annotation["tags"] == {"stage": "carol"}
    assert annotation["qc_reviewers"] == ["carol"]
    assert annotation["qc_history"][0]["undone_by"] == "bob"
    assert annotation["qc_history"][0]["undone_at"] is not None

    annotation = store.undo_quality_check(task["id"], 0, "carol")

    assert annotation["mos"] == 3
    assert annotation["tags"] == {"stage": "annotated"}
    assert annotation["qc_reviewers"] == []


def test_quality_check_api_saves_and_undoes_review_edits():
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "a.jpg", "dst_image": "b.jpg"}])

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
        subtask = annotations_app.store.assign_subtask(task["id"], "alice")
        annotations_app.store.save_annotation(task["id"], subtask["id"], 0, "alice", 3, {"stage": "annotated"})
        client = annotations_app.app.test_client()

        save_response = client.post(
            f"/api/tasks/{task['id']}/results/0/qc",
            json={"username": "bob", "mos": 5, "tags": {"stage": "checked"}},
        )

        assert save_response.status_code == 200
        saved = save_response.get_json()["annotation"]
        assert saved["mos"] == 5
        assert saved["tags"] == {"stage": "checked"}
        assert saved["qc_reviewers"] == ["bob"]

        undo_response = client.delete(
            f"/api/tasks/{task['id']}/results/0/qc",
            json={"username": "bob"},
        )

        assert undo_response.status_code == 200
        undone = undo_response.get_json()["annotation"]
        assert undone["mos"] == 3
        assert undone["tags"] == {"stage": "annotated"}
        assert undone["qc_reviewers"] == []
    finally:
        annotations_app.store = old_store


def test_result_tag_click_handler_opens_editor_from_entire_row():
    script = (PROJECT_ROOT / "web" / "annotations" / "static" / "app.js").read_text(encoding="utf-8")
    handler_start = script.index('$("resultTags").addEventListener("click"')
    handler = script[handler_start:script.index('  $("prevBtn").addEventListener', handler_start)]

    assert 'event.target.closest(".qcInlineEditor")' in handler
    assert 'event.target.closest(".resultTagValue")' not in handler
    assert "beginResultTagEdit(row)" in handler


def test_results_view_supports_configured_next_page_shortcut():
    script = (PROJECT_ROOT / "web" / "annotations" / "static" / "app.js").read_text(encoding="utf-8")
    keydown_start = script.index('document.addEventListener("keydown"')
    keydown_handler = script[keydown_start:script.index("  });\n}", keydown_start)]
    results_branch = keydown_handler[
        keydown_handler.index('if (!$("resultsView").classList.contains("hidden"))')
        : keydown_handler.index('if ($("annotateView").classList.contains("hidden"))')
    ]

    assert 'event.key.toUpperCase() === state.nextKey.toUpperCase()' in results_branch
    assert "state.resultPage += 1" in results_branch
    assert "renderResultPage()" in results_branch


def test_annotation_view_combines_mos_and_pager_controls():
    template = (PROJECT_ROOT / "web" / "annotations" / "templates" / "index.html").read_text(encoding="utf-8")
    annotate_start = template.index('<section id="annotateView"')
    annotate_end = template.index('<section id="resultsView"', annotate_start)
    annotate_markup = template[annotate_start:annotate_end]
    controls_start = annotate_markup.index('<div class="annotationControls"')
    review_grid_start = annotate_markup.index('<div class="reviewGrid">')
    review_grid_end = annotate_markup.index('</section>', review_grid_start)

    assert review_grid_start < controls_start < review_grid_end
    assert controls_start < annotate_markup.index('<div class="scoreBar" id="scoreBar"></div>')
    assert controls_start < annotate_markup.index('<div class="pager annotationPager">')


def test_annotation_view_uses_narrow_scrollable_tag_panel_without_heading():
    template = (PROJECT_ROOT / "web" / "annotations" / "templates" / "index.html").read_text(encoding="utf-8")
    styles = (PROJECT_ROOT / "web" / "annotations" / "static" / "styles.css").read_text(encoding="utf-8")
    annotate_start = template.index('<section id="annotateView"')
    annotate_end = template.index('<section id="resultsView"', annotate_start)
    annotate_markup = template[annotate_start:annotate_end]
    tag_editor_start = annotate_markup.index('<aside class="tagEditor">')
    tag_editor_end = annotate_markup.index('</aside>', tag_editor_start)
    tag_editor_markup = annotate_markup[tag_editor_start:tag_editor_end]
    tag_editor_styles = styles[
        styles.index("#annotateView .tagEditor {"):styles.index("#annotateView .labelGroup {")
    ]

    assert "<h3>标签</h3>" not in tag_editor_markup
    assert "--review-panel-height: min(80vh, calc(100vh - 190px));" in styles
    assert "grid-template-columns: minmax(240px, 1fr) minmax(240px, 1fr) minmax(360px, 440px);" in styles
    assert "height: var(--review-panel-height);" in tag_editor_styles
    assert "overflow: auto;" in tag_editor_styles
    assert "max-height: none;" in tag_editor_styles


def test_annotation_view_exposes_abandon_subtask_action():
    template = (PROJECT_ROOT / "web" / "annotations" / "templates" / "index.html").read_text(encoding="utf-8")
    script = (PROJECT_ROOT / "web" / "annotations" / "static" / "app.js").read_text(encoding="utf-8")

    assert 'id="abandonSubtaskBtn"' in template
    assert "abandonSubtask()" in script
    assert "method: \"DELETE\"" in script
    assert "/subtasks/${state.subtask.id}" in script


def test_create_task_form_exposes_shuffle_before_split_option():
    template = (PROJECT_ROOT / "web" / "annotations" / "templates" / "index.html").read_text(encoding="utf-8")
    script = (PROJECT_ROOT / "web" / "annotations" / "static" / "app.js").read_text(encoding="utf-8")

    assert 'id="shuffleItemsInput"' in template
    assert "随机打乱后切分" in template
    assert 'shuffle_items: $("shuffleItemsInput").checked' in script


def test_pie_chart_legend_groups_percent_and_count_with_label():
    script = (PROJECT_ROOT / "web" / "annotations" / "static" / "app.js").read_text(encoding="utf-8")
    styles = (PROJECT_ROOT / "web" / "annotations" / "static" / "styles.css").read_text(encoding="utf-8")

    assert "formatStatsLegendValue(item.count, total)" in script
    assert "<strong>${item.count}</strong>" not in script
    assert "12px minmax(90px, 1fr) auto minmax(58px, auto)" not in styles


def test_statistics_counts_filtered_dimensions_and_combinations():
    from web.annotations.app import AnnotationStore
    from web.annotations.label_options import LABEL_OPTION_GROUPS

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [{"src_image": f"src/{idx}.jpg", "dst_image": f"dst/{idx}.jpg"} for idx in range(4)],
    )
    input_group = LABEL_OPTION_GROUPS[0]["name"]
    angle_name = LABEL_OPTION_GROUPS[0]["dimensions"][0]["name"]
    food_name = LABEL_OPTION_GROUPS[0]["dimensions"][1]["name"]
    output_group = LABEL_OPTION_GROUPS[1]["name"]
    shot_method_name = LABEL_OPTION_GROUPS[1]["dimensions"][2]["name"]

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=4)
    alice_subtask = store.assign_subtask(task["id"], "alice")
    rows = [
        ("alice", 5, 0, "food-a", "method-a"),
        ("alice", 5, 45, "food-a", "method-b"),
        ("alice", 4, 75, "food-b", "method-a"),
        ("alice", 3, 90, "food-b", "method-b"),
    ]
    for item_index, (username, mos, angle, food, method) in enumerate(rows):
        store.save_annotation(
            task["id"],
            alice_subtask["id"],
            item_index=item_index,
            username=username,
            mos=mos,
            tags={
                input_group: {angle_name: angle, food_name: food},
                output_group: {shot_method_name: method},
            },
        )

    statistics = store.get_statistics(
        task["id"],
        filters={"labels": [{"path": [input_group, angle_name], "ranges": [{"min": 0, "max": 60}]}]},
        combinations=[
            [{"type": "mos"}, {"type": "label", "path": [output_group, shot_method_name]}],
            [
                {"type": "mos"},
                {"type": "label", "path": [input_group, food_name]},
                {"type": "label", "path": [output_group, shot_method_name]},
            ],
        ],
    )

    assert statistics["total"] == 2
    assert statistics["annotators"]["items"] == [{"label": "alice", "count": 2}]
    assert statistics["mos"]["items"] == [{"label": "5", "count": 2}]

    label_stats = {
        tuple(item["path"]): item["items"]
        for item in statistics["labels"]
    }
    assert label_stats[(input_group, angle_name)] == [{"label": "0", "count": 1}, {"label": "45", "count": 1}]
    assert label_stats[(input_group, food_name)] == [{"label": "food-a", "count": 2}]

    assert statistics["combinations"][0]["items"] == [
        {"label": "MOS 5 + method-a", "count": 1},
        {"label": "MOS 5 + method-b", "count": 1},
    ]
    assert statistics["combinations"][1]["items"] == [
        {"label": "MOS 5 + food-a + method-a", "count": 1},
        {"label": "MOS 5 + food-a + method-b", "count": 1},
    ]


def test_refresh_task_labels_updates_ai_labels_without_overwriting_user_annotations():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    annotation_dir = tmp_path / "labels"
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "src/a.jpg", "dst_image": "dst/a.jpg"}])

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1, annotation_dir=str(annotation_dir))
    subtask = store.assign_subtask(task["id"], "alice")
    store.save_annotation(task["id"], subtask["id"], 0, "alice", 5, {"manual": "keep"})

    write_json(annotation_dir / "dst" / "a.json", {"labels": {"ai_score": 4}})

    result = store.refresh_task_labels(task["id"])
    refreshed_task = store.get_task(task["id"])
    refreshed_subtask = store.get_subtask(task["id"], subtask["id"])

    assert result["updated_count"] == 1
    refreshed_labels = refreshed_subtask["items"][0]["labels"]
    assert list(refreshed_labels.values()) == [{"ai_score": 4}]
    assert refreshed_subtask["items"][0]["annotation"]["tags"] == {"manual": "keep"}
    assert refreshed_task["annotation_count"] == 1
    assert json.loads((Path(refreshed_task["data_dir"]) / "annotations" / "0.json").read_text(encoding="utf-8"))[
        "tags"
    ] == {"manual": "keep"}


def test_image_endpoint_limits_long_edge_by_default_and_can_return_original():
    from PIL import Image
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    make_test_image(tmp_path / "large.jpg", (2048, 512))
    make_test_image(tmp_path / "dst.jpg", (64, 64))
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "large.jpg", "dst_image": "dst.jpg"}])

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
        client = annotations_app.app.test_client()

        preview_response = client.get(f"/api/tasks/{task['id']}/images/0/src")
        original_response = client.get(f"/api/tasks/{task['id']}/images/0/src?original=1")

        preview = Image.open(BytesIO(preview_response.data))
        original = Image.open(BytesIO(original_response.data))
        assert max(preview.size) == 1024
        assert original.size == (2048, 512)
    finally:
        annotations_app.store = old_store


def test_image_endpoint_caches_resized_preview_in_task_data_dir():
    from PIL import Image
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    make_test_image(tmp_path / "large.jpg", (2048, 512))
    make_test_image(tmp_path / "dst.jpg", (64, 64))
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "large.jpg", "dst_image": "dst.jpg"}])

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
        client = annotations_app.app.test_client()

        preview_response = client.get(f"/api/tasks/{task['id']}/images/0/src")

        assert preview_response.status_code == 200
        preview = Image.open(BytesIO(preview_response.data))
        assert max(preview.size) == 1024
        cached_files = list((Path(task["data_dir"]) / "preview_cache").glob("*.jpg"))
        assert len(cached_files) == 1
        assert cached_files[0].stat().st_size == len(preview_response.data)
    finally:
        annotations_app.store = old_store


def test_resized_image_file_ignores_incomplete_tmp_preview_cache_files():
    from web.annotations.app import preview_cache_key, resized_image_file

    tmp_path = make_workspace_tmp()
    image_path = tmp_path / "large.jpg"
    cache_dir = tmp_path / "preview_cache"
    make_test_image(image_path, (2048, 512))
    cache_dir.mkdir()
    cache_key = preview_cache_key(image_path)
    incomplete_tmp_path = cache_dir / f"{cache_key}.jpg.12345.tmp"
    incomplete_tmp_path.write_bytes(b"incomplete image bytes")

    preview_path, mimetype = resized_image_file(image_path, cache_dir)

    assert preview_path.name == f"{cache_key}.jpg"
    assert preview_path.exists()
    assert preview_path.read_bytes() != incomplete_tmp_path.read_bytes()
    assert mimetype == "image/jpeg"


def test_image_endpoint_uses_configured_preview_cache_dir():
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    custom_cache_dir = tmp_path / "custom-preview-cache"
    make_test_image(tmp_path / "large.jpg", (2048, 512))
    make_test_image(tmp_path / "dst.jpg", (64, 64))
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "large.jpg", "dst_image": "dst.jpg"}])

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json", preview_cache_dir=custom_cache_dir)
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
        client = annotations_app.app.test_client()

        response = client.get(f"/api/tasks/{task['id']}/images/0/src")

        assert response.status_code == 200
        cached_files = list((custom_cache_dir / task["id"]).glob("*.jpg"))
        assert len(cached_files) == 1
        assert not (Path(task["data_dir"]) / "preview_cache").exists()
    finally:
        annotations_app.store = old_store


def test_preview_cache_job_generates_all_input_image_previews():
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore, PreviewCacheJobs

    tmp_path = make_workspace_tmp()
    make_test_image(tmp_path / "src" / "0.jpg", (1400, 800))
    make_test_image(tmp_path / "src" / "1.jpg", (1200, 900))
    make_test_image(tmp_path / "dst" / "0.jpg", (64, 64))
    make_test_image(tmp_path / "dst" / "1.jpg", (64, 64))
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [
            {"src_image": "src/0.jpg", "dst_image": "dst/0.jpg"},
            {"src_image": "src/1.jpg", "dst_image": "dst/1.jpg"},
        ],
    )

    old_store = annotations_app.store
    old_jobs = annotations_app.preview_cache_jobs
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.preview_cache_jobs = PreviewCacheJobs(annotations_app.store)
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=2)
        client = annotations_app.app.test_client()

        start_response = client.post(f"/api/tasks/{task['id']}/preview-cache/jobs")

        assert start_response.status_code == 202
        job_id = start_response.get_json()["job"]["id"]
        for _ in range(50):
            job_response = client.get(f"/api/tasks/{task['id']}/preview-cache/jobs/{job_id}")
            job = job_response.get_json()["job"]
            if job["status"] == "completed":
                break
            time.sleep(0.02)
        else:
            raise AssertionError("preview cache job did not complete")

        assert job["progress"] == 100
        assert job["result"]["total"] == 2
        assert job["result"]["generated_count"] == 2
        assert job["result"]["failed_count"] == 0
        assert len(list((Path(task["data_dir"]) / "preview_cache").glob("*.jpg"))) == 2
    finally:
        annotations_app.preview_cache_jobs = old_jobs
        annotations_app.store = old_store


def test_task_list_exposes_input_preview_cache_action_after_visualization():
    script = (PROJECT_ROOT / "web" / "annotations" / "static" / "app.js").read_text(encoding="utf-8")
    render_start = script.index("function renderTasks()")
    render_end = script.index("async function deleteTask", render_start)
    render_tasks = script[render_start:render_end]

    assert render_tasks.index('data-action="visualization"') < render_tasks.index('data-action="cache-inputs"')
    assert "缓存输入图" in render_tasks
    assert "warmInputPreviewCache(taskId, button)" in script


def test_task_list_only_renders_delete_action_for_admin_user_at_the_end():
    script = (PROJECT_ROOT / "web" / "annotations" / "static" / "app.js").read_text(encoding="utf-8")
    render_start = script.index("function renderTasks()")
    render_end = script.index("async function deleteTask", render_start)
    render_tasks = script[render_start:render_end]

    assert 'state.username === "孙本猿"' in render_tasks
    assert render_tasks.index('data-action="cache-inputs"') < render_tasks.index('data-action="delete"')
    assert render_tasks.rindex('data-action="delete"') > render_tasks.rindex('data-action="cache-inputs"')


def test_issue_ui_exposes_task_result_and_detail_actions():
    template = (PROJECT_ROOT / "web" / "annotations" / "templates" / "index.html").read_text(encoding="utf-8")
    script = (PROJECT_ROOT / "web" / "annotations" / "static" / "app.js").read_text(encoding="utf-8")

    assert '"issuesView"' in script
    assert 'id="issuesView"' in template
    assert 'data-action="issues"' in script
    assert 'id="createIssueBtn"' in template
    assert 'id="issueModal"' in template
    assert 'id="issueAnswerInput"' in script
    assert 'openIssues(taskId, taskName)' in script
    assert 'openIssueModal()' in script
    assert 'submitIssueAnswer()' in script


def test_issue_ui_supports_markdown_export_and_bbox_insertion():
    script = (PROJECT_ROOT / "web" / "annotations" / "static" / "app.js").read_text(encoding="utf-8")
    styles = (PROJECT_ROOT / "web" / "annotations" / "static" / "styles.css").read_text(encoding="utf-8")

    assert '/issues/export.md' in script
    assert 'function formatBboxReference' in script
    assert 'function beginIssueRegionSelection' in script
    assert '[${imageKind}: x=${' in script
    assert 'issueImageSelection' in styles
    assert 'issueModalCard' in styles


def test_preview_cache_dir_can_be_configured_by_environment(monkeypatch):
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    custom_cache_dir = tmp_path / "env-preview-cache"
    monkeypatch.setenv("ANNOTATIONS_PREVIEW_CACHE_DIR", str(custom_cache_dir))

    store = AnnotationStore(tmp_path / "state.json")

    assert store.preview_cache_dir("task-1") == custom_cache_dir / "task-1"


def test_image_path_reads_item_from_chunk_storage_without_items_json():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    write_jsonl(
        tmp_path / "data.jsonl",
        [
            {"src_image": f"src/{idx}.jpg", "dst_image": f"dst/{idx}.jpg"}
            for idx in range(3)
        ],
    )

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(tmp_path / "data.jsonl"), chunk_size=2)
    data_dir = Path(task["data_dir"])

    assert not (data_dir / "items.json").exists()
    assert store.image_path(task["id"], 2, "dst") == tmp_path / "dst" / "2.jpg"


def test_downloads_export_all_annotated_results_as_jsonl_and_xlsx():
    from openpyxl import load_workbook
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(
        jsonl_path,
        [
            {"src_image": "a.jpg", "dst_image": "b.jpg"},
            {"src_image": "c.jpg", "dst_image": "d.jpg"},
        ],
    )

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=2)
        subtask = annotations_app.store.assign_subtask(task["id"], "alice")
        annotations_app.store.save_annotation(task["id"], subtask["id"], 0, "alice", 2, {"bad": True})
        annotations_app.store.save_annotation(task["id"], subtask["id"], 1, "alice", 5, {"good": True})

        client = annotations_app.app.test_client()
        jsonl_response = client.get(f"/api/tasks/{task['id']}/download?format=jsonl")
        xlsx_response = client.get(f"/api/tasks/{task['id']}/download?format=xlsx")

        rows = [json.loads(line) for line in jsonl_response.data.decode("utf-8").splitlines()]
        assert [row["item_index"] for row in rows] == [0, 1]
        assert rows[0]["mos"] == 2
        assert jsonl_response.headers["Content-Disposition"].endswith('filename=food_annotations.jsonl')

        workbook = load_workbook(BytesIO(xlsx_response.data))
        sheet = workbook.active
        assert sheet.max_row == 3
        assert sheet["A1"].value == "item_index"
        assert sheet["A2"].value == 0
        assert sheet["F3"].value == 5
    finally:
        annotations_app.store = old_store


def test_issue_creation_assigns_annotator_and_snapshots_result():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "src/a.jpg", "dst_image": "dst/a.jpg"}])

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
    subtask = store.assign_subtask(task["id"], "alice")
    store.save_annotation(task["id"], subtask["id"], 0, "alice", 3, {"quality": "needs review"})

    issue = store.create_issue(
        task["id"],
        item_index=0,
        created_by="reviewer",
        title="Please check score",
        body="MOS looks too low",
    )

    assert issue["status"] == "open"
    assert issue["created_by"] == "reviewer"
    assert issue["assigned_to"] == "alice"
    assert issue["item_index"] == 0
    assert issue["snapshot"]["mos"] == 3
    assert issue["snapshot"]["tags"] == {"quality": "needs review"}
    assert issue["snapshot"]["src_image"] == "src/a.jpg"
    assert issue["snapshot"]["dst_image"] == "dst/a.jpg"
    assert store.list_issues(task["id"])[0]["id"] == issue["id"]


def test_issue_answers_status_changes_and_markdown_export():
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "src/a.jpg", "dst_image": "dst/a.jpg"}])

    store = AnnotationStore(tmp_path / "state.json")
    task = store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
    subtask = store.assign_subtask(task["id"], "alice")
    store.save_annotation(task["id"], subtask["id"], 0, "alice", 4, {"dish": "noodle"})
    issue = store.create_issue(task["id"], 0, "reviewer", "", "")

    answered = store.add_issue_answer(
        task["id"],
        issue["id"],
        "alice",
        "Region is correct [dst: x=0.100 y=0.200 w=0.300 h=0.400]",
    )
    closed = store.close_issue(task["id"], issue["id"], "reviewer")
    reopened = store.reopen_issue(task["id"], issue["id"], "reviewer")
    markdown = store.export_issues_markdown(task["id"])

    assert answered["answers"][0]["author"] == "alice"
    assert answered["answers"][0]["body"] == "Region is correct [dst: x=0.100 y=0.200 w=0.300 h=0.400]"
    assert closed["status"] == "closed"
    assert closed["closed_by"] == "reviewer"
    assert reopened["status"] == "open"
    assert reopened["closed_at"] is None
    assert "# Issues for food" in markdown
    assert "Assignee: alice" in markdown
    assert "src/a.jpg" in markdown
    assert "Region is correct [dst: x=0.100 y=0.200 w=0.300 h=0.400]" in markdown


def test_issue_api_endpoints_create_answer_close_reopen_and_export_markdown():
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "src/a.jpg", "dst_image": "dst/a.jpg"}])

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
        subtask = annotations_app.store.assign_subtask(task["id"], "alice")
        annotations_app.store.save_annotation(task["id"], subtask["id"], 0, "alice", 5, {"ok": True})
        client = annotations_app.app.test_client()

        create_response = client.post(
            f"/api/tasks/{task['id']}/issues",
            json={"item_index": 0, "created_by": "reviewer", "title": "", "body": "Please explain"},
        )
        issue = create_response.get_json()["issue"]
        answer_response = client.post(
            f"/api/tasks/{task['id']}/issues/{issue['id']}/answers",
            json={"author": "alice", "body": "Because [src: x=0.000 y=0.100 w=0.200 h=0.300]"},
        )
        close_response = client.post(
            f"/api/tasks/{task['id']}/issues/{issue['id']}/close",
            json={"username": "reviewer"},
        )
        reopen_response = client.post(
            f"/api/tasks/{task['id']}/issues/{issue['id']}/reopen",
            json={"username": "reviewer"},
        )
        export_response = client.get(f"/api/tasks/{task['id']}/issues/export.md")

        assert create_response.status_code == 201
        assert issue["title"] == "请检查该条标注结果"
        assert issue["assigned_to"] == "alice"
        assert answer_response.get_json()["issue"]["answers"][0]["author"] == "alice"
        assert close_response.get_json()["issue"]["status"] == "closed"
        assert reopen_response.get_json()["issue"]["status"] == "open"
        assert export_response.mimetype == "text/markdown"
        assert "Please explain" in export_response.data.decode("utf-8")
    finally:
        annotations_app.store = old_store


def test_jsonl_download_with_non_ascii_task_name_returns_over_http():
    from werkzeug.serving import make_server
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "a.jpg", "dst_image": "b.jpg"}])

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    server = None
    try:
        task = annotations_app.store.create_task("美食数据", str(tmp_path), str(jsonl_path), chunk_size=1)
        subtask = annotations_app.store.assign_subtask(task["id"], "alice")
        annotations_app.store.save_annotation(task["id"], subtask["id"], 0, "alice", 5, {"good": True})

        server = make_server("127.0.0.1", 0, annotations_app.app)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        port = server.server_port

        with urllib.request.urlopen(
            f"http://127.0.0.1:{port}/api/tasks/{task['id']}/download?format=jsonl",
            timeout=5,
        ) as response:
            body = response.read().decode("utf-8")

        assert response.status == 200
        assert json.loads(body)["item_index"] == 0
        assert "filename*=" in response.headers["Content-Disposition"]
    finally:
        if server is not None:
            server.shutdown()
        annotations_app.store = old_store


def test_delete_task_api_requires_admin_user_and_reports_missing_task():
    from web.annotations import app as annotations_app
    from web.annotations.app import AnnotationStore

    tmp_path = make_workspace_tmp()
    jsonl_path = tmp_path / "data.jsonl"
    write_jsonl(jsonl_path, [{"src_image": "a.jpg", "dst_image": "b.jpg"}])

    old_store = annotations_app.store
    annotations_app.store = AnnotationStore(tmp_path / "state.json")
    annotations_app.app.config.update(TESTING=True)
    try:
        task = annotations_app.store.create_task("food", str(tmp_path), str(jsonl_path), chunk_size=1)
        client = annotations_app.app.test_client()

        forbidden = client.delete(f"/api/tasks/{task['id']}", json={"username": "alice"})
        response = client.delete(f"/api/tasks/{task['id']}", json={"username": "孙本猿"})
        missing = client.delete(f"/api/tasks/{task['id']}", json={"username": "孙本猿"})

        assert forbidden.status_code == 403
        assert response.status_code == 200
        assert annotations_app.store.get_task(task["id"]) is None
        assert annotations_app.store.list_tasks() == []
        assert missing.status_code == 404
    finally:
        annotations_app.store = old_store
